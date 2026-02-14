from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from reporting.batch_excel import DEFAULT_BATCH_EXCEL_FILE, generate_batch_summary_excel


def test_generate_batch_summary_excel_writes_all_sheets(tmp_path: Path) -> None:
    summary = {
        "output_dir_abs": str(tmp_path),
        "items": [
            {
                "relative_path": "a/paper.pdf",
                "status": "success",
                "run_id": "run_1",
                "runtime_ms": 1200,
                "overall_risk": "low",
                "domain_risks": {
                    "D1": "low",
                    "D2": "low",
                    "D3": "low",
                    "D4": "low",
                    "D5": "low",
                },
                "error": None,
            },
            {
                "relative_path": "b/fail.pdf",
                "status": "failed",
                "run_id": "",
                "runtime_ms": None,
                "overall_risk": None,
                "domain_risks": {},
                "error": "ValueError: boom",
            },
        ],
    }
    summary_path = tmp_path / "batch_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    result_path = tmp_path / "a" / "paper" / "result.json"
    result_path.parent.mkdir(parents=True, exist_ok=True)
    result_path.write_text(
        json.dumps(
            {
                "run_id": "run_1",
                "result": {
                    "variant": "standard",
                    "question_set_version": "rob2-standard-1.0",
                    "overall": {"risk": "low", "rationale": "ok"},
                    "domains": [
                        {
                            "domain": "D1",
                            "effect_type": None,
                            "risk": "low",
                            "risk_rationale": "low risk rationale",
                            "answers": [
                                {
                                    "question_id": "q1_1",
                                    "rob2_id": "q1_1",
                                    "text": "Was the allocation sequence random?",
                                    "answer": "Y",
                                    "confidence": 0.9,
                                    "rationale": "randomized",
                                    "evidence_refs": [
                                        {
                                            "paragraph_id": "p1-0001",
                                            "page": 1,
                                            "title": "title",
                                            "quote": "randomized text",
                                        }
                                    ],
                                }
                            ],
                            "missing_questions": ["q1_2"],
                            "rule_trace": ["D1:R1 q1_1 in YES -> low"],
                        }
                    ],
                    "citations": [
                        {
                            "paragraph_id": "p1-0001",
                            "page": 1,
                            "title": "title",
                            "text": "full text",
                            "uses": [
                                {
                                    "domain": "D1",
                                    "question_id": "q1_1",
                                    "quote": "randomized text",
                                }
                            ],
                        }
                    ],
                    "document_metadata": {
                        "title": "Sample Trial",
                        "authors": ["Alice", "Bob"],
                        "year": 2020,
                        "affiliations": ["Hospital A"],
                        "funders": ["Fund A"],
                        "sources": [
                            {
                                "paragraph_id": "p1-0001",
                                "quote": "Sample Trial",
                            }
                        ],
                        "extraction": {
                            "method": "langextract",
                            "model_id": "model-x",
                            "provider": "anthropic",
                            "confidence": 0.88,
                            "error": None,
                        },
                    },
                },
                "audit_reports": [
                    {
                        "domain": "D1",
                        "audited_questions": 3,
                        "mismatches": [
                            {
                                "question_id": "q1_1",
                                "domain": "D1",
                                "effect_type": None,
                                "domain_answer": "Y",
                                "audit_answer": "PY",
                                "audit_confidence": 0.72,
                                "audit_evidence": [
                                    {"paragraph_id": "p1-0001", "quote": "randomized text"}
                                ],
                            }
                        ],
                        "patch_window": 1,
                        "patches_applied": {"q1_1": 1},
                        "rerun_enabled": True,
                        "domain_rerun": True,
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    output = tmp_path / DEFAULT_BATCH_EXCEL_FILE
    exported = generate_batch_summary_excel(summary, summary_path=summary_path, output_path=output)

    assert exported == 2
    assert output.exists()

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "00_批次总览",
        "01_域风险",
        "02_决策路径",
        "03_问答明细",
        "04_证据明细",
        "05_文献元信息",
        "06_元信息来源",
        "07_审计差异",
        "08_引用汇总",
    ]

    overview_sheet = workbook["00_批次总览"]
    overview_header = [cell.value for cell in overview_sheet[1]]
    assert overview_header == [
        "relative_path",
        "status",
        "run_id",
        "runtime_seconds",
        "overall_risk",
        "D1_risk",
        "D2_risk",
        "D3_risk",
        "D4_risk",
        "D5_risk",
        "error",
        "result_json_path",
    ]
    assert overview_sheet[2][3].value == 1.2
    assert overview_sheet.max_row == 3

    domain_sheet = workbook["01_域风险"]
    domain_paths = {
        str(row[0].value)
        for row in domain_sheet.iter_rows(min_row=2, max_row=domain_sheet.max_row)
    }
    assert "a/paper.pdf" in domain_paths
    assert "b/fail.pdf" not in domain_paths

    metadata_sheet = workbook["05_文献元信息"]
    metadata_row = [cell.value for cell in metadata_sheet[2]]
    assert metadata_row[2] == "Sample Trial"
    assert metadata_row[3] == "Alice; Bob"
    assert metadata_row[5] == "Hospital A"

    audit_sheet = workbook["07_审计差异"]
    assert audit_sheet.max_row >= 2
    assert audit_sheet[2][2].value == "D1"
    assert audit_sheet[2][5].value == "q1_1"


def test_generate_batch_summary_excel_keeps_running_when_result_invalid(tmp_path: Path) -> None:
    summary = {
        "output_dir_abs": str(tmp_path),
        "items": [
            {
                "relative_path": "paper.pdf",
                "status": "success",
                "run_id": "run_x",
                "runtime_ms": 10,
                "overall_risk": "low",
                "domain_risks": {"D1": "low"},
                "error": "",
            }
        ],
    }
    summary_path = tmp_path / "batch_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    result_path = tmp_path / "paper" / "result.json"
    result_path.parent.mkdir(parents=True, exist_ok=True)
    result_path.write_text("{bad json}", encoding="utf-8")

    output = tmp_path / DEFAULT_BATCH_EXCEL_FILE
    generate_batch_summary_excel(summary, summary_path=summary_path, output_path=output)

    workbook = load_workbook(output)
    overview_sheet = workbook["00_批次总览"]
    overview_row = [cell.value for cell in overview_sheet[2]]
    assert "result.json 读取失败" in str(overview_row[10])

    domain_sheet = workbook["01_域风险"]
    assert domain_sheet.max_row == 1
