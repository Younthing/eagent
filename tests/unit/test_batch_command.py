from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
import typer
from openpyxl import load_workbook

from cli.commands import batch as batch_command


def test_discover_pdfs_recursively_sorted(tmp_path: Path) -> None:
    (tmp_path / "b").mkdir()
    (tmp_path / "a").mkdir()
    (tmp_path / "a" / "x.PDF").write_bytes(b"%PDF-1.4")
    (tmp_path / "b" / "y.pdf").write_bytes(b"%PDF-1.4")
    (tmp_path / "z.txt").write_text("ignore", encoding="utf-8")

    files = batch_command._discover_pdfs(tmp_path)
    rel = [item.relative_to(tmp_path).as_posix() for item in files]

    assert rel == ["a/x.PDF", "b/y.pdf"]


def test_checkpoint_compatibility_requires_reset() -> None:
    checkpoint = {
        "input_dir_abs": "/input",
        "output_dir_abs": "/out",
        "options_hash": "old",
        "file_list_hash": "old_file_list",
        "batch_id": None,
        "batch_name": None,
        "items": [],
    }

    with pytest.raises(typer.BadParameter):
        batch_command._assert_checkpoint_compatible(
            checkpoint,
            input_dir_abs="/input",
            output_dir_abs="/out",
            file_list_hash="new_file_list",
            batch_id=None,
            batch_name=None,
        )


def test_checkpoint_compatibility_ignores_options_hash() -> None:
    checkpoint = {
        "input_dir_abs": "/input",
        "output_dir_abs": "/out",
        "options_hash": "old",
        "file_list_hash": "same",
        "batch_id": None,
        "batch_name": None,
        "items": [],
    }

    batch_command._assert_checkpoint_compatible(
        checkpoint,
        input_dir_abs="/input",
        output_dir_abs="/out",
        file_list_hash="same",
        batch_id=None,
        batch_name=None,
    )


def test_checkpoint_compatibility_ignores_batch_name_when_batch_id_provided() -> None:
    checkpoint = {
        "input_dir_abs": "/input",
        "output_dir_abs": "/out",
        "options_hash": "same",
        "file_list_hash": "same",
        "batch_id": "batch_1",
        "batch_name": "old_name",
        "items": [],
    }

    batch_command._assert_checkpoint_compatible(
        checkpoint,
        input_dir_abs="/input",
        output_dir_abs="/out",
        file_list_hash="same",
        batch_id="batch_1",
        batch_name="new_name",
    )


def test_file_list_hash_changes_when_pdf_content_changes(tmp_path: Path) -> None:
    pdf = tmp_path / "paper.pdf"
    pdf.write_bytes(b"%PDF-1.4\nfirst")
    files = batch_command._discover_pdfs(tmp_path)
    first_entries = batch_command._build_file_entries(tmp_path, files)
    first_hash = batch_command._build_file_list_hash(first_entries)

    pdf.write_bytes(b"%PDF-1.4\nsecond")
    files = batch_command._discover_pdfs(tmp_path)
    second_entries = batch_command._build_file_entries(tmp_path, files)
    second_hash = batch_command._build_file_list_hash(second_entries)

    assert first_hash != second_hash


def test_build_reusable_result_index_uses_batch_meta(tmp_path: Path) -> None:
    reusable_dir = tmp_path / "cache" / "x"
    reusable_dir.mkdir(parents=True)
    (reusable_dir / "result.json").write_text(
        json.dumps(
            {
                "run_id": "run_1",
                "runtime_ms": 100,
                "result": {"overall": {"risk": "low"}, "domains": []},
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (reusable_dir / "batch_item_meta.json").write_text(
        json.dumps(
            {
                "version": 1,
                "pdf_sha256": "hash_abc",
                "relative_path": "foo.pdf",
                "updated_at": "2026-01-01T00:00:00+00:00",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    index = batch_command._build_reusable_result_index(tmp_path)
    assert index == {"hash_abc": reusable_dir.resolve()}


def test_materialize_reused_output_copies_managed_files(tmp_path: Path) -> None:
    source_dir = tmp_path / "source"
    target_dir = tmp_path / "target"
    source_dir.mkdir()
    target_dir.mkdir()

    (source_dir / "result.json").write_text("{}", encoding="utf-8")
    (source_dir / "report.html").write_text("<html/>", encoding="utf-8")
    (target_dir / "result.json").write_text('{"old":true}', encoding="utf-8")
    (target_dir / "report.pdf").write_text("old", encoding="utf-8")

    batch_command._materialize_reused_output(source_dir=source_dir, target_dir=target_dir)

    assert (target_dir / "result.json").read_text(encoding="utf-8") == "{}"
    assert (target_dir / "report.html").read_text(encoding="utf-8") == "<html/>"
    assert not (target_dir / "report.pdf").exists()


def test_load_checkpoint_requires_v2_and_pdf_hash(tmp_path: Path) -> None:
    checkpoint_path = tmp_path / "batch_checkpoint.json"
    checkpoint_path.write_text(
        json.dumps(
            {
                "version": 1,
                "items": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    with pytest.raises(typer.BadParameter):
        batch_command._load_checkpoint(checkpoint_path)

    checkpoint_path.write_text(
        json.dumps(
            {
                "version": 2,
                "items": [{"relative_path": "a.pdf"}],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    with pytest.raises(typer.BadParameter):
        batch_command._load_checkpoint(checkpoint_path)


def test_write_summary_files_contains_expected_columns(tmp_path: Path) -> None:
    checkpoint = {
        "version": 1,
        "input_dir_abs": "/input",
        "output_dir_abs": "/out",
        "batch_id": "batch_1",
        "batch_name": "nightly",
        "items": [
            {
                "relative_path": "a/paper.pdf",
                "output_subdir": "a/paper",
                "status": "success",
                "run_id": "run_1",
                "runtime_ms": 123,
                "overall_risk": "low",
                "domain_risks": {
                    "D1": "low",
                    "D2": "low",
                    "D3": "some concerns",
                    "D4": "low",
                    "D5": "low",
                },
                "error": None,
                "updated_at": "2026-01-01T00:00:00+00:00",
            }
        ],
    }

    batch_command._write_summary_files(checkpoint, tmp_path)

    summary = json.loads((tmp_path / "batch_summary.json").read_text(encoding="utf-8"))
    assert summary["counts"]["success"] == 1

    with (tmp_path / "batch_summary.csv").open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    assert reader.fieldnames == [
        "relative_path",
        "status",
        "run_id",
        "runtime_ms",
        "overall_risk",
        "D1_risk",
        "D2_risk",
        "D3_risk",
        "D4_risk",
        "D5_risk",
        "error",
    ]
    assert len(rows) == 1
    assert rows[0]["relative_path"] == "a/paper.pdf"
    assert rows[0]["D3_risk"] == "some concerns"


def test_plot_batch_accepts_directory_source(tmp_path: Path) -> None:
    batch_dir = tmp_path / "batch"
    batch_dir.mkdir()
    summary = {
        "items": [
            {
                "relative_path": "a/paper.pdf",
                "status": "success",
                "overall_risk": "low",
                "domain_risks": {
                    "D1": "low",
                    "D2": "some concerns",
                    "D3": "high",
                    "D4": "low",
                    "D5": "low",
                },
            }
        ]
    }
    (batch_dir / "batch_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    batch_command.plot_batch(
        source=batch_dir,
        output=None,
        include_non_success=False,
    )

    image_path = batch_dir / "batch_traffic_light.png"
    assert image_path.exists()
    assert image_path.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")


def test_plot_batch_accepts_summary_file_source(tmp_path: Path) -> None:
    summary_path = tmp_path / "batch_summary.json"
    summary = {
        "items": [
            {
                "relative_path": "a/paper.pdf",
                "status": "success",
                "overall_risk": "high",
                "domain_risks": {
                    "D1": "high",
                    "D2": "some_concerns",
                    "D3": "low",
                    "D4": "low",
                    "D5": "low",
                },
            }
        ]
    }
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    output = tmp_path / "custom_plot.png"
    batch_command.plot_batch(
        source=summary_path,
        output=output,
        include_non_success=False,
    )

    assert output.exists()
    assert output.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")


def test_plot_batch_raises_when_no_plot_rows(tmp_path: Path) -> None:
    summary_path = tmp_path / "batch_summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "relative_path": "a/paper.pdf",
                        "status": "failed",
                        "overall_risk": None,
                        "domain_risks": {},
                    }
                ]
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    with pytest.raises(typer.BadParameter, match="红绿灯图生成失败"):
        batch_command.plot_batch(
            source=summary_path,
            output=None,
            include_non_success=False,
        )


def test_excel_batch_accepts_directory_source(tmp_path: Path) -> None:
    batch_dir = tmp_path / "batch"
    batch_dir.mkdir()
    summary = {
        "output_dir_abs": str(batch_dir),
        "items": [
            {
                "relative_path": "a/paper.pdf",
                "status": "success",
                "run_id": "run_1",
                "runtime_ms": 123,
                "overall_risk": "low",
                "domain_risks": {"D1": "low"},
                "error": None,
            }
        ],
    }
    (batch_dir / "batch_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    result_path = batch_dir / "a" / "paper" / "result.json"
    result_path.parent.mkdir(parents=True, exist_ok=True)
    result_path.write_text(
        json.dumps(
            {
                "run_id": "run_1",
                "result": {
                    "variant": "standard",
                    "question_set_version": "1.0",
                    "overall": {"risk": "low", "rationale": "ok"},
                    "domains": [],
                    "citations": [],
                    "document_metadata": None,
                },
                "audit_reports": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    batch_command.excel_batch(source=batch_dir, output=None)

    excel_path = batch_dir / "batch_summary.xlsx"
    assert excel_path.exists()
    workbook = load_workbook(excel_path)
    assert "00_批次总览" in workbook.sheetnames


def test_excel_batch_accepts_summary_file_source(tmp_path: Path) -> None:
    summary_path = tmp_path / "batch_summary.json"
    summary = {
        "output_dir_abs": str(tmp_path),
        "items": [
            {
                "relative_path": "paper.pdf",
                "status": "failed",
                "run_id": None,
                "runtime_ms": None,
                "overall_risk": None,
                "domain_risks": {},
                "error": "boom",
            }
        ],
    }
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    output = tmp_path / "custom_summary.xlsx"
    batch_command.excel_batch(source=summary_path, output=output)

    assert output.exists()
    workbook = load_workbook(output)
    assert "00_批次总览" in workbook.sheetnames


def test_retryable_error_detection() -> None:
    assert batch_command._is_retryable_error("RuntimeError: 429 rate limit")
    assert batch_command._is_retryable_error("TimeoutError: request timeout")
    assert batch_command._is_retryable_error("ConnectTimeout: upstream timeout")
    assert not batch_command._is_retryable_error("ValueError: bad input")


def test_adaptive_concurrency_controller_down_and_up() -> None:
    controller = batch_command._AdaptiveConcurrencyController(
        mode="adaptive",
        current_limit=3,
        min_limit=1,
        max_limit=4,
        success_window=2,
    )

    controller.observe(success=True, had_retryable_error=True)
    assert controller.current_limit == 2

    controller.observe(success=True, had_retryable_error=False)
    assert controller.current_limit == 2
    controller.observe(success=True, had_retryable_error=False)
    assert controller.current_limit == 3


def test_fixed_concurrency_controller_keeps_limit() -> None:
    controller = batch_command._AdaptiveConcurrencyController(
        mode="fixed",
        current_limit=3,
        min_limit=1,
        max_limit=4,
        success_window=2,
    )
    controller.observe(success=False, had_retryable_error=True)
    controller.observe(success=True, had_retryable_error=False)
    controller.observe(success=True, had_retryable_error=False)
    assert controller.current_limit == 3
