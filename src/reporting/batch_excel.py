"""Batch Excel exporter for ROB2 summaries."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_BATCH_EXCEL_FILE = "batch_summary.xlsx"

_SUCCESS_STATUSES = {"success", "skipped"}

_SHEET_COLUMNS: dict[str, list[str]] = {
    "00_批次总览": [
        "relative_path",
        "status",
        "run_id",
        "runtime_ms",
        "overall_risk",
        "D1_risk",
        "D2_risk",
        "D3_risk",
        "D4_risk",
        "D5_risk",
        "error",
        "result_json_path",
    ],
    "01_域风险": [
        "relative_path",
        "run_id",
        "domain",
        "effect_type",
        "domain_risk",
        "risk_rationale",
        "missing_questions_count",
        "missing_questions",
    ],
    "02_决策路径": [
        "relative_path",
        "run_id",
        "domain",
        "rule_trace_index",
        "rule_trace",
        "missing_questions",
    ],
    "03_问答明细": [
        "relative_path",
        "run_id",
        "domain",
        "question_id",
        "rob2_id",
        "question_text",
        "answer",
        "confidence",
        "rationale",
        "evidence_ref_count",
    ],
    "04_证据明细": [
        "relative_path",
        "run_id",
        "domain",
        "question_id",
        "evidence_index",
        "paragraph_id",
        "page",
        "title",
        "quote",
    ],
    "05_文献元信息": [
        "relative_path",
        "run_id",
        "title",
        "authors",
        "year",
        "affiliations",
        "funders",
        "extraction_method",
        "extraction_model_id",
        "extraction_provider",
        "extraction_confidence",
        "extraction_error",
    ],
    "06_元信息来源": [
        "relative_path",
        "run_id",
        "source_index",
        "paragraph_id",
        "quote",
    ],
    "07_审计差异": [
        "relative_path",
        "run_id",
        "audit_domain",
        "audited_questions",
        "mismatch_index",
        "question_id",
        "domain",
        "effect_type",
        "domain_answer",
        "audit_answer",
        "audit_confidence",
        "audit_evidence_count",
        "patch_window",
        "patches_applied",
        "rerun_enabled",
        "domain_rerun",
    ],
    "08_引用汇总": [
        "relative_path",
        "run_id",
        "paragraph_id",
        "page",
        "title",
        "use_index",
        "use_domain",
        "use_question_id",
        "use_quote",
    ],
}

_SHEET_ORDER = list(_SHEET_COLUMNS.keys())


def generate_batch_summary_excel(
    summary: Mapping[str, Any],
    *,
    summary_path: Path,
    output_path: Path,
) -> int:
    """Generate the batch workbook from summary + per-item result.json files."""
    items = summary.get("items")
    if not isinstance(items, list):
        raise ValueError("summary items 缺失")

    batch_output_dir = _resolve_batch_output_dir(summary, summary_path)
    rows_by_sheet = _build_rows(items=items, batch_output_dir=batch_output_dir)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in _SHEET_ORDER:
        sheet = workbook.create_sheet(title=sheet_name)
        columns = _SHEET_COLUMNS[sheet_name]
        _write_sheet(sheet, columns=columns, rows=rows_by_sheet[sheet_name])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(rows_by_sheet["00_批次总览"])


def _resolve_batch_output_dir(summary: Mapping[str, Any], summary_path: Path) -> Path:
    value = summary.get("output_dir_abs")
    if isinstance(value, str) and value.strip():
        return Path(value).resolve()
    return summary_path.parent.resolve()


def _build_rows(
    *,
    items: list[Any],
    batch_output_dir: Path,
) -> dict[str, list[dict[str, Any]]]:
    rows_by_sheet = {name: [] for name in _SHEET_ORDER}
    overview_rows = rows_by_sheet["00_批次总览"]
    domain_rows = rows_by_sheet["01_域风险"]
    trace_rows = rows_by_sheet["02_决策路径"]
    answer_rows = rows_by_sheet["03_问答明细"]
    evidence_rows = rows_by_sheet["04_证据明细"]
    metadata_rows = rows_by_sheet["05_文献元信息"]
    metadata_source_rows = rows_by_sheet["06_元信息来源"]
    audit_rows = rows_by_sheet["07_审计差异"]
    citation_rows = rows_by_sheet["08_引用汇总"]

    for raw_item in items:
        if not isinstance(raw_item, Mapping):
            continue

        relative_path = _to_str(raw_item.get("relative_path"))
        status = _to_str(raw_item.get("status")).strip().lower()
        run_id = _to_str(raw_item.get("run_id"))
        runtime_ms = _empty_if_none(raw_item.get("runtime_ms"))
        overall_risk = _to_str(raw_item.get("overall_risk"))
        domain_risks = raw_item.get("domain_risks")
        if not isinstance(domain_risks, Mapping):
            domain_risks = {}
        error = _to_str(raw_item.get("error"))

        result_json_path = _build_result_json_path(batch_output_dir, relative_path)
        parsed: _ParsedResult | None = None

        if status in _SUCCESS_STATUSES:
            parsed, parse_error = _load_result_payload(result_json_path)
            if parse_error:
                error = _append_error(error, f"result.json 读取失败: {parse_error}")

        overview_rows.append(
            {
                "relative_path": relative_path,
                "status": status,
                "run_id": run_id,
                "runtime_ms": runtime_ms,
                "overall_risk": overall_risk,
                "D1_risk": _to_str(domain_risks.get("D1")),
                "D2_risk": _to_str(domain_risks.get("D2")),
                "D3_risk": _to_str(domain_risks.get("D3")),
                "D4_risk": _to_str(domain_risks.get("D4")),
                "D5_risk": _to_str(domain_risks.get("D5")),
                "error": error,
                "result_json_path": str(result_json_path),
            }
        )

        if parsed is None:
            continue

        effective_run_id = parsed.run_id or run_id
        result_obj = parsed.result

        domains = result_obj.get("domains")
        if not isinstance(domains, list):
            domains = []

        for domain in domains:
            if not isinstance(domain, Mapping):
                continue
            domain_id = _to_str(domain.get("domain"))
            effect_type = _to_str(domain.get("effect_type"))
            domain_risk = _to_str(domain.get("risk"))
            risk_rationale = _to_str(domain.get("risk_rationale"))

            missing_questions = domain.get("missing_questions")
            if not isinstance(missing_questions, list):
                missing_questions = []
            missing_joined = _join_values(missing_questions)

            domain_rows.append(
                {
                    "relative_path": relative_path,
                    "run_id": effective_run_id,
                    "domain": domain_id,
                    "effect_type": effect_type,
                    "domain_risk": domain_risk,
                    "risk_rationale": risk_rationale,
                    "missing_questions_count": len(missing_questions),
                    "missing_questions": missing_joined,
                }
            )

            rule_trace = domain.get("rule_trace")
            if not isinstance(rule_trace, list):
                rule_trace = []
            if rule_trace:
                for index, trace in enumerate(rule_trace, start=1):
                    trace_rows.append(
                        {
                            "relative_path": relative_path,
                            "run_id": effective_run_id,
                            "domain": domain_id,
                            "rule_trace_index": index,
                            "rule_trace": _to_str(trace),
                            "missing_questions": missing_joined,
                        }
                    )
            else:
                trace_rows.append(
                    {
                        "relative_path": relative_path,
                        "run_id": effective_run_id,
                        "domain": domain_id,
                        "rule_trace_index": "",
                        "rule_trace": "",
                        "missing_questions": missing_joined,
                    }
                )

            answers = domain.get("answers")
            if not isinstance(answers, list):
                answers = []
            for answer in answers:
                if not isinstance(answer, Mapping):
                    continue

                question_id = _to_str(answer.get("question_id"))
                evidence_refs = answer.get("evidence_refs")
                if not isinstance(evidence_refs, list):
                    evidence_refs = []

                answer_rows.append(
                    {
                        "relative_path": relative_path,
                        "run_id": effective_run_id,
                        "domain": domain_id,
                        "question_id": question_id,
                        "rob2_id": _to_str(answer.get("rob2_id")),
                        "question_text": _to_str(answer.get("text")),
                        "answer": _to_str(answer.get("answer")),
                        "confidence": _empty_if_none(answer.get("confidence")),
                        "rationale": _to_str(answer.get("rationale")),
                        "evidence_ref_count": len(evidence_refs),
                    }
                )

                for evidence_index, ref in enumerate(evidence_refs, start=1):
                    if not isinstance(ref, Mapping):
                        continue
                    evidence_rows.append(
                        {
                            "relative_path": relative_path,
                            "run_id": effective_run_id,
                            "domain": domain_id,
                            "question_id": question_id,
                            "evidence_index": evidence_index,
                            "paragraph_id": _to_str(ref.get("paragraph_id")),
                            "page": _empty_if_none(ref.get("page")),
                            "title": _to_str(ref.get("title")),
                            "quote": _to_str(ref.get("quote")),
                        }
                    )

        metadata = result_obj.get("document_metadata")
        metadata_map = metadata if isinstance(metadata, Mapping) else {}
        extraction = metadata_map.get("extraction")
        if not isinstance(extraction, Mapping):
            extraction = {}

        metadata_rows.append(
            {
                "relative_path": relative_path,
                "run_id": effective_run_id,
                "title": _to_str(metadata_map.get("title")),
                "authors": _join_values(metadata_map.get("authors")),
                "year": _empty_if_none(metadata_map.get("year")),
                "affiliations": _join_values(metadata_map.get("affiliations")),
                "funders": _join_values(metadata_map.get("funders")),
                "extraction_method": _to_str(extraction.get("method")),
                "extraction_model_id": _to_str(extraction.get("model_id")),
                "extraction_provider": _to_str(extraction.get("provider")),
                "extraction_confidence": _empty_if_none(extraction.get("confidence")),
                "extraction_error": _to_str(extraction.get("error")),
            }
        )

        sources = metadata_map.get("sources")
        if isinstance(sources, list):
            for source_index, source in enumerate(sources, start=1):
                if not isinstance(source, Mapping):
                    continue
                metadata_source_rows.append(
                    {
                        "relative_path": relative_path,
                        "run_id": effective_run_id,
                        "source_index": source_index,
                        "paragraph_id": _to_str(source.get("paragraph_id")),
                        "quote": _to_str(source.get("quote")),
                    }
                )

        _append_audit_rows(
            audit_rows,
            relative_path=relative_path,
            run_id=effective_run_id,
            audit_reports=parsed.audit_reports,
        )
        _append_citation_rows(
            citation_rows,
            relative_path=relative_path,
            run_id=effective_run_id,
            citations=result_obj.get("citations"),
        )

    return rows_by_sheet


def _append_audit_rows(
    rows: list[dict[str, Any]],
    *,
    relative_path: str,
    run_id: str,
    audit_reports: Any,
) -> None:
    if not isinstance(audit_reports, list):
        return

    for report in audit_reports:
        if not isinstance(report, Mapping):
            continue

        base = {
            "relative_path": relative_path,
            "run_id": run_id,
            "audit_domain": _to_str(report.get("domain")),
            "audited_questions": _empty_if_none(report.get("audited_questions")),
            "patch_window": _empty_if_none(report.get("patch_window")),
            "patches_applied": _stringify_patches_applied(report.get("patches_applied")),
            "rerun_enabled": _empty_if_none(report.get("rerun_enabled")),
            "domain_rerun": _empty_if_none(report.get("domain_rerun")),
        }

        mismatches = report.get("mismatches")
        if not isinstance(mismatches, list):
            mismatches = []

        if not mismatches:
            rows.append(
                {
                    **base,
                    "mismatch_index": "",
                    "question_id": "",
                    "domain": "",
                    "effect_type": "",
                    "domain_answer": "",
                    "audit_answer": "",
                    "audit_confidence": "",
                    "audit_evidence_count": "",
                }
            )
            continue

        for mismatch_index, mismatch in enumerate(mismatches, start=1):
            if not isinstance(mismatch, Mapping):
                continue
            audit_evidence = mismatch.get("audit_evidence")
            evidence_count = len(audit_evidence) if isinstance(audit_evidence, list) else 0
            rows.append(
                {
                    **base,
                    "mismatch_index": mismatch_index,
                    "question_id": _to_str(mismatch.get("question_id")),
                    "domain": _to_str(mismatch.get("domain")),
                    "effect_type": _to_str(mismatch.get("effect_type")),
                    "domain_answer": _to_str(mismatch.get("domain_answer")),
                    "audit_answer": _to_str(mismatch.get("audit_answer")),
                    "audit_confidence": _empty_if_none(mismatch.get("audit_confidence")),
                    "audit_evidence_count": evidence_count,
                }
            )


def _append_citation_rows(
    rows: list[dict[str, Any]],
    *,
    relative_path: str,
    run_id: str,
    citations: Any,
) -> None:
    if not isinstance(citations, list):
        return

    for citation in citations:
        if not isinstance(citation, Mapping):
            continue

        base = {
            "relative_path": relative_path,
            "run_id": run_id,
            "paragraph_id": _to_str(citation.get("paragraph_id")),
            "page": _empty_if_none(citation.get("page")),
            "title": _to_str(citation.get("title")),
        }

        uses = citation.get("uses")
        if not isinstance(uses, list):
            uses = []

        if not uses:
            rows.append(
                {
                    **base,
                    "use_index": "",
                    "use_domain": "",
                    "use_question_id": "",
                    "use_quote": "",
                }
            )
            continue

        for use_index, use in enumerate(uses, start=1):
            if not isinstance(use, Mapping):
                continue
            rows.append(
                {
                    **base,
                    "use_index": use_index,
                    "use_domain": _to_str(use.get("domain")),
                    "use_question_id": _to_str(use.get("question_id")),
                    "use_quote": _to_str(use.get("quote")),
                }
            )


def _write_sheet(
    sheet: Worksheet,
    *,
    columns: Iterable[str],
    rows: list[dict[str, Any]],
) -> None:
    column_list = list(columns)
    sheet.append(column_list)
    sheet.freeze_panes = "A2"

    for row in rows:
        values = [_empty_if_none(row.get(column)) for column in column_list]
        sheet.append(values)

    if sheet.max_column > 0:
        sheet.auto_filter.ref = sheet.dimensions


def _build_result_json_path(batch_output_dir: Path, relative_path: str) -> Path:
    if not relative_path:
        return batch_output_dir / "result.json"
    subdir = Path(relative_path).with_suffix("")
    return (batch_output_dir / subdir / "result.json").resolve()


class _ParsedResult:
    def __init__(self, *, run_id: str, result: Mapping[str, Any], audit_reports: Any) -> None:
        self.run_id = run_id
        self.result = result
        self.audit_reports = audit_reports


def _load_result_payload(path: Path) -> tuple[_ParsedResult | None, str | None]:
    if not path.exists():
        return None, "文件不存在"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return None, str(exc)
    if not isinstance(payload, Mapping):
        return None, "result.json 顶层不是对象"

    result = payload.get("result")
    if not isinstance(result, Mapping):
        return None, "缺少 result 对象"

    run_id = _to_str(payload.get("run_id"))
    return (
        _ParsedResult(run_id=run_id, result=result, audit_reports=payload.get("audit_reports")),
        None,
    )


def _append_error(current: str, extra: str) -> str:
    cur = current.strip()
    ext = extra.strip()
    if not cur:
        return ext
    if not ext:
        return cur
    return f"{cur} | {ext}"


def _stringify_patches_applied(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Mapping):
        parts = []
        for key in sorted(value.keys(), key=lambda item: str(item)):
            parts.append(f"{key}:{value[key]}")
        return "; ".join(parts)
    return _to_str(value)


def _join_values(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    parts: list[str] = []
    for item in value:
        text = _to_str(item).strip()
        if text:
            parts.append(text)
    return "; ".join(parts)


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _empty_if_none(value: Any) -> Any:
    if value is None:
        return ""
    return value


__all__ = ["DEFAULT_BATCH_EXCEL_FILE", "generate_batch_summary_excel"]
